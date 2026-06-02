"""
blueprints/exports.py — Excel/EDI/PDF 匯出報表（出勤、薪資、請假、加班、員工、訓練、績效、費用、扣繳）
"""
import io
from datetime import datetime as _dt, date as _date
from flask import Blueprint, request, jsonify, Response

from auth import login_required, require_module
from config import TW_TZ
from db import get_db

bp = Blueprint('exports', __name__)

# ── PDF 輔助函式 ───────────────────────────────────────────────────

def _register_zh_font():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    try:
        pdfmetrics.getFont('STSong-Light')
    except Exception:
        pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))


def _pdf_response(buf, filename):
    buf.seek(0)
    return Response(buf.read(), mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename={filename}'})


def _build_pdf(title, subtitle, headers, col_widths, rows, landscape=False):
    """產生帶標題與表格的 PDF，回傳 BytesIO。"""
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.pagesizes import A4, landscape as LS
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    _register_zh_font()
    FONT = 'STSong-Light'
    buf = io.BytesIO()
    pagesize = LS(A4) if landscape else A4
    doc = SimpleDocTemplate(buf, pagesize=pagesize,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    title_style = ParagraphStyle('title', fontName=FONT, fontSize=14, spaceAfter=4)
    sub_style   = ParagraphStyle('sub',   fontName=FONT, fontSize=9,  textColor=colors.grey, spaceAfter=8)

    total_w = sum(col_widths)
    page_w  = (pagesize[0] - 3*cm)
    scale   = page_w / total_w if total_w > page_w else 1.0
    scaled  = [w * scale for w in col_widths]

    table_data = [headers] + [[str(v) if v is not None else '' for v in r] for r in rows]
    t = Table(table_data, colWidths=scaled, repeatRows=1)
    HDR_BG = colors.HexColor('#0F1C3A')
    ROW_BG = colors.HexColor('#F4F6FA')
    t.setStyle(TableStyle([
        ('FONT',        (0, 0), (-1, -1), FONT, 8),
        ('FONT',        (0, 0), (-1,  0), FONT, 9),
        ('BACKGROUND',  (0, 0), (-1,  0), HDR_BG),
        ('TEXTCOLOR',   (0, 0), (-1,  0), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ROW_BG]),
        ('GRID',        (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ALIGN',       (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',  (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 3),
        ('WORDWRAP',    (0, 0), (-1, -1), True),
    ]))

    story = [Paragraph(title, title_style)]
    if subtitle:
        story.append(Paragraph(subtitle, sub_style))
    story.append(t)
    doc.build(story)
    return buf

# ── Excel 輔助函式（含完整格式） ───────────────────────────────────

def _xl_workbook(sheet_name='Sheet1'):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    return wb, ws


def _xl_write_header(ws, headers, col_widths):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin',  color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
    )
    hdr_fill = PatternFill('solid', fgColor='0F1C3A')
    hdr_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = thin
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def _xl_write_rows(ws, data_rows, num_cols, number_cols=None):
    from openpyxl.styles import Alignment, PatternFill, Border, Side
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin',  color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
    )
    even_fill   = PatternFill('solid', fgColor='F4F6FA')
    center      = Alignment(horizontal='center', vertical='center')
    left_align  = Alignment(horizontal='left',   vertical='center')
    number_cols = set(number_cols or [])
    for ri, row_vals in enumerate(data_rows, 2):
        fill = even_fill if ri % 2 == 0 else None
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if fill: cell.fill = fill
            cell.border = thin
            if ci in number_cols and isinstance(v, (int, float)):
                cell.alignment = center
                cell.number_format = '#,##0.00'
            else:
                cell.alignment = center if isinstance(v, (int, float, type(None))) else left_align


def _xl_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'})


# ── EDI 輔助函式 ───────────────────────────────────────────────────

def _get_insurance_settings():
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT setting_key, setting_value FROM insurance_settings").fetchall()
        return {r['setting_key']: r['setting_value'] for r in rows}
    except Exception:
        return {}


def _roc_date(date_str):
    if not date_str: return '0000000'
    try:
        d = _date.fromisoformat(str(date_str)[:10])
        return f'{d.year - 1911:03d}{d.month:02d}{d.day:02d}'
    except Exception:
        return '0000000'


def _edi_bytes(val, width, numeric=False):
    s = str(val or '')
    if numeric:
        return s.rjust(width, '0').encode('ascii', errors='replace')[:width]
    try: b = s.encode('big5', errors='replace')
    except: b = s.encode('ascii', errors='replace')
    b = b + b' ' * max(0, width - len(b))
    return b[:width]


def _get_edi_staff(staff_ids_str):
    with get_db() as conn:
        if staff_ids_str:
            ids = [int(x) for x in staff_ids_str.split(',') if x.strip().isdigit()]
            return conn.execute(
                "SELECT * FROM punch_staff WHERE id = ANY(%s) AND active=TRUE ORDER BY name", (ids,)
            ).fetchall()
        return conn.execute("SELECT * FROM punch_staff WHERE active=TRUE ORDER BY name").fetchall()


# ── Attendance Export ──────────────────────────────────────────────

@bp.route('/api/export/attendance', methods=['GET'])
@login_required
def api_export_attendance():
    month    = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    staff_id = request.args.get('staff_id', '')
    conds, params = ["TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s"], [month]
    if staff_id:
        conds.append("pr.staff_id=%s"); params.append(int(staff_id))
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT ps.employee_code, ps.name as staff_name, ps.department, ps.role,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   pr.punch_type,
                   to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei', 'HH24:MI') as punch_time,
                   pr.is_manual, pr.gps_distance, pr.location_name, pr.note
            FROM punch_records pr JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE {' AND '.join(conds)} ORDER BY ps.name, pr.punched_at
        """, params).fetchall()
    PUNCH_LABEL = {'in': '上班打卡', 'out': '下班打卡', 'break_out': '休息開始', 'break_in': '休息結束'}
    wb, ws = _xl_workbook(f'{month} 出勤明細')
    headers = ['員工代碼', '姓名', '部門', '職稱', '日期', '打卡類型', '時間', '補打卡', 'GPS距離(m)', '地點', '備註']
    widths  = [10, 10, 12, 12, 12, 10, 8, 7, 11, 14, 20]
    _xl_write_header(ws, headers, widths)
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['staff_name'], r['department'] or '', r['role'] or '',
         str(r['work_date']), PUNCH_LABEL.get(r['punch_type'], r['punch_type']),
         r['punch_time'], '是' if r['is_manual'] else '',
         float(r['gps_distance']) if r['gps_distance'] is not None else '',
         r['location_name'] or '', r['note'] or '']
        for r in rows
    ], len(headers))
    return _xl_response(wb, f'attendance_{month}.xlsx')


@bp.route('/api/export/attendance-summary', methods=['GET'])
@login_required
def api_export_attendance_summary():
    month = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT ps.employee_code, ps.name, ps.department, ps.role,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_out,
                   MIN(CASE WHEN pr.punch_type='in'  THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as ci_ts,
                   MAX(CASE WHEN pr.punch_type='out' THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as co_ts,
                   BOOL_OR(pr.is_manual) as has_manual, COUNT(*) as punch_count
            FROM punch_records pr JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY ps.employee_code, ps.name, ps.department, ps.role,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY ps.name, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
        """, (month,)).fetchall()
    wb, ws = _xl_workbook(f'{month} 出勤摘要')
    headers = ['員工代碼', '姓名', '部門', '職稱', '日期', '上班', '下班', '工時(h)', '打卡次數', '含補打']
    widths  = [10, 10, 12, 12, 12, 8, 8, 9, 9, 7]
    _xl_write_header(ws, headers, widths)
    data = []
    for r in rows:
        dur_h = ''
        if r['ci_ts'] and r['co_ts']:
            try:
                ci = r['ci_ts'] if hasattr(r['ci_ts'], 'timestamp') else _dt.fromisoformat(str(r['ci_ts']))
                co = r['co_ts'] if hasattr(r['co_ts'], 'timestamp') else _dt.fromisoformat(str(r['co_ts']))
                dur_h = round((co - ci).total_seconds() / 3600, 2)
            except Exception:
                pass
        data.append([r['employee_code'] or '', r['name'], r['department'] or '', r['role'] or '',
                     str(r['work_date']), r['clock_in'] or '', r['clock_out'] or '',
                     dur_h, r['punch_count'], '是' if r['has_manual'] else ''])
    _xl_write_rows(ws, data, len(headers), number_cols={8, 9})
    return _xl_response(wb, f'attendance_summary_{month}.xlsx')


@bp.route('/api/attendance/anomaly-report', methods=['GET'])
@login_required
def api_anomaly_report_excel():
    import openpyxl, calendar as _cal
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    month = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    try:
        y, mo = int(month[:4]), int(month[5:7])
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400

    with get_db() as conn:
        punch_rows = conn.execute("""
            SELECT ps.id as staff_id, ps.name as staff_name, ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN (pr.punched_at AT TIME ZONE 'Asia/Taipei') END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN (pr.punched_at AT TIME ZONE 'Asia/Taipei') END) as clock_out,
                   BOOL_OR(pr.punch_type='in')  as has_in,
                   BOOL_OR(pr.punch_type='out') as has_out
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id AND ps.active=TRUE
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY ps.id, ps.name, ps.department, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY work_date, ps.name
        """, (month,)).fetchall()
        shift_rows = conn.execute("""
            SELECT sa.staff_id, sa.shift_date,
                   st.start_time::text as start_time, st.end_time::text as end_time,
                   ps.name as staff_name, ps.department
            FROM shift_assignments sa JOIN shift_types st ON st.id=sa.shift_type_id
            JOIN punch_staff ps ON ps.id=sa.staff_id AND ps.active=TRUE
            WHERE TO_CHAR(sa.shift_date,'YYYY-MM')=%s
        """, (month,)).fetchall()
        first_day = f"{y}-{mo:02d}-01"
        last_day  = f"{y}-{mo:02d}-{_cal.monthrange(y, mo)[1]:02d}"
        leave_rows = conn.execute("""
            SELECT staff_id, start_date, end_date FROM leave_requests
            WHERE status='approved' AND start_date <= %s AND end_date >= %s
        """, (last_day, first_day)).fetchall()

    shift_map = {(r['staff_id'], str(r['shift_date'])): r for r in shift_rows}
    leave_set = set()
    for lr in leave_rows:
        s = _date.fromisoformat(str(lr['start_date']))
        e = _date.fromisoformat(str(lr['end_date']))
        cur = s
        from datetime import timedelta as _td
        while cur <= e:
            leave_set.add((lr['staff_id'], str(cur)))
            cur += _td(days=1)

    today = _date.today()
    anomalies = []
    for r in punch_rows:
        ds = str(r['work_date']); sid = r['staff_id']
        shift = shift_map.get((sid, ds))
        anomaly_type = ''; detail = ''
        if not r['has_in'] and r['has_out']:
            anomaly_type = '缺上班打卡'; detail = f"僅有下班 {str(r['clock_out'])[11:16]}"
        elif r['has_in'] and not r['has_out']:
            if _date.fromisoformat(ds) < today:
                anomaly_type = '缺下班打卡'; detail = f"上班 {str(r['clock_in'])[11:16]} 無下班"
        elif r['has_in'] and r['has_out'] and shift:
            ci_t = str(r['clock_in'])[11:16]; sh_s = str(shift['start_time'])[:5]
            try:
                late_mins = (int(ci_t[:2])*60 + int(ci_t[3:5])) - (int(sh_s[:2])*60 + int(sh_s[3:5]))
                if late_mins > 10:
                    anomaly_type = '遲到'; detail = f"應 {sh_s}，實際 {ci_t}（+{late_mins}分）"
            except Exception: pass
            if not anomaly_type:
                co_t = str(r['clock_out'])[11:16]; sh_e = str(shift['end_time'])[:5]
                try:
                    early_mins = (int(sh_e[:2])*60 + int(sh_e[3:5])) - (int(co_t[:2])*60 + int(co_t[3:5]))
                    if early_mins > 15:
                        anomaly_type = '早退'; detail = f"應 {sh_e}，實際 {co_t}（-{early_mins}分）"
                except Exception: pass
        if anomaly_type:
            anomalies.append({
                'staff_name': r['staff_name'], 'department': r['department'] or '', 'date': ds,
                'shift_start': str(shift['start_time'])[:5] if shift else '—',
                'shift_end':   str(shift['end_time'])[:5]   if shift else '—',
                'clock_in':    str(r['clock_in'])[11:16]  if r['clock_in']  else '—',
                'clock_out':   str(r['clock_out'])[11:16] if r['clock_out'] else '—',
                'anomaly_type': anomaly_type, 'detail': detail,
            })
    punched_set = {(r['staff_id'], str(r['work_date'])) for r in punch_rows}
    for sr in shift_rows:
        ds = str(sr['shift_date']); sid = sr['staff_id']
        if _date.fromisoformat(ds) >= today: continue
        if (sid, ds) in punched_set or (sid, ds) in leave_set: continue
        anomalies.append({
            'staff_name': sr['staff_name'], 'department': sr['department'] or '', 'date': ds,
            'shift_start': str(sr['start_time'])[:5], 'shift_end': str(sr['end_time'])[:5],
            'clock_in': '—', 'clock_out': '—', 'anomaly_type': '未打卡',
            'detail': f"排班 {str(sr['start_time'])[:5]}～{str(sr['end_time'])[:5]}，整日無打卡記錄",
        })
    anomalies.sort(key=lambda x: (x['date'], x['staff_name']))

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = f'{month} 異常明細'
    thin = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                  top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
    header_fill = PatternFill('solid', fgColor='0F1C3A')
    warn_fill   = PatternFill('solid', fgColor='FFF3CD')
    err_fill    = PatternFill('solid', fgColor='FDECEA')
    center_align = Alignment(horizontal='center', vertical='center')
    headers = ['員工姓名', '部門', '日期', '應上班', '應下班', '實際上班', '實際下班', '異常類型', '說明']
    col_w   = [12, 10, 12, 8, 8, 8, 8, 12, 30]
    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Noto Sans TC', size=11)
        cell.fill = header_fill; cell.alignment = center_align; cell.border = thin
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w
    for ri, a in enumerate(anomalies, 2):
        row_fill = err_fill if a['anomaly_type'] in ('缺上班打卡', '缺下班打卡') else warn_fill
        for ci, v in enumerate([a['staff_name'], a['department'], a['date'],
                                  a['shift_start'], a['shift_end'], a['clock_in'], a['clock_out'],
                                  a['anomaly_type'], a['detail']], 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.fill = row_fill; cell.border = thin
            cell.alignment = center_align if ci != 9 else Alignment(vertical='center')
    ws.row_dimensions[1].height = 22; ws.freeze_panes = 'A2'
    ws2 = wb.create_sheet('摘要'); ws2.append(['統計', '數量']); ws2.append(['異常總筆數', len(anomalies)])
    by_type = {}
    for a in anomalies: by_type[a['anomaly_type']] = by_type.get(a['anomaly_type'], 0) + 1
    for t, c in sorted(by_type.items(), key=lambda x: -x[1]): ws2.append([t, c])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=anomaly_{month}.xlsx'})


# ── Salary Export ──────────────────────────────────────────────────

@bp.route('/api/export/salary', methods=['GET'])
@login_required
def api_export_salary():
    month = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT sr.*, ps.name as staff_name, ps.employee_code, ps.department, ps.role, ps.salary_type
            FROM salary_records sr JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.month = %s ORDER BY ps.name
        """, (month,)).fetchall()
        leave_detail = conn.execute("""
            SELECT lr.staff_id,
                   COALESCE(SUM(CASE WHEN lt.code='personal' THEN lr.total_days ELSE 0 END), 0) AS personal_days,
                   COALESCE(SUM(CASE WHEN lt.code='sick'     THEN lr.total_days ELSE 0 END), 0) AS sick_days
            FROM leave_requests lr JOIN leave_types lt ON lt.id = lr.leave_type_id
            WHERE lr.status='approved' AND to_char(lr.start_date,'YYYY-MM')=%s GROUP BY lr.staff_id
        """, (month,)).fetchall()
    leave_map = {r['staff_id']: r for r in leave_detail}
    wb, ws = _xl_workbook(f'{month} 薪資明細')
    headers = ['員工代碼', '姓名', '部門', '職稱', '薪資制度',
               '工作日', '出勤天數', '請假天數', '無薪假天數', '事假天數', '病假天數',
               '津貼合計', '扣除合計', '加班費', '實領金額', '狀態', '備註']
    widths  = [10, 10, 12, 12, 8, 8, 8, 8, 9, 8, 8, 11, 11, 10, 12, 8, 20]
    _xl_write_header(ws, headers, widths)
    data = []
    for r in rows:
        ld = leave_map.get(r['staff_id'])
        data.append([r['employee_code'] or '', r['staff_name'], r['department'] or '', r['role'] or '',
                     '時薪制' if r['salary_type'] == 'hourly' else '月薪制',
                     float(r['work_days'] or 0), float(r['actual_days'] or 0),
                     float(r['leave_days'] or 0), float(r['unpaid_days'] or 0),
                     float(ld['personal_days'] if ld else 0), float(ld['sick_days'] if ld else 0),
                     float(r['allowance_total'] or 0), float(r['deduction_total'] or 0),
                     float(r['ot_pay'] or 0), float(r['net_pay'] or 0),
                     '已確認' if r['status'] == 'confirmed' else '草稿', r['note'] or ''])
    _xl_write_rows(ws, data, len(headers), number_cols={6,7,8,9,10,11,12,13,14,15})
    return _xl_response(wb, f'salary_{month}.xlsx')


# ── Leave Export ───────────────────────────────────────────────────

@bp.route('/api/export/leave', methods=['GET'])
@login_required
def api_export_leave():
    month = request.args.get('month', ''); year = request.args.get('year', '')
    staff_id = request.args.get('staff_id', ''); status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if status:   conds.append("lr.status=%s");                               params.append(status)
    if month:    conds.append("to_char(lr.start_date,'YYYY-MM')=%s");        params.append(month)
    if year:     conds.append("EXTRACT(YEAR FROM lr.start_date)=%s");         params.append(int(year))
    if staff_id: conds.append("lr.staff_id=%s");                             params.append(int(staff_id))
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT lr.*, ps.name as staff_name, ps.employee_code, ps.department,
                   lt.name as leave_type_name, lt.pay_rate
            FROM leave_requests lr JOIN punch_staff ps ON ps.id = lr.staff_id
            JOIN leave_types lt ON lt.id = lr.leave_type_id
            WHERE {' AND '.join(conds)} ORDER BY lr.start_date, ps.name
        """, params).fetchall()
    PAY_LABEL = {1.0: '全薪', 0.5: '半薪', 0.0: '無薪'}
    STATUS_LABEL = {'approved': '已核准', 'rejected': '已退回', 'pending': '待審核'}
    wb, ws = _xl_workbook('請假記錄')
    headers = ['員工代碼', '姓名', '部門', '假別', '薪資倍率', '開始日期', '結束日期', '天數', '原因', '代理人', '狀態']
    widths  = [10, 10, 12, 10, 8, 12, 12, 7, 24, 10, 8]
    _xl_write_header(ws, headers, widths)
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['staff_name'], r['department'] or '',
         r['leave_type_name'], PAY_LABEL.get(float(r['pay_rate']), f"{r['pay_rate']}倍"),
         str(r['start_date']), str(r['end_date']), float(r['total_days']),
         r['reason'] or '', r['substitute_name'] or '', STATUS_LABEL.get(r['status'], r['status'])]
        for r in rows
    ], len(headers), number_cols={8})
    return _xl_response(wb, f'leave_{month or year or "all"}.xlsx')


# ── Overtime Export ────────────────────────────────────────────────

@bp.route('/api/export/overtime', methods=['GET'])
@login_required
def api_export_overtime():
    month = request.args.get('month', ''); staff_id = request.args.get('staff_id', '')
    status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if month:    conds.append("to_char(r.request_date,'YYYY-MM')=%s"); params.append(month)
    if staff_id: conds.append("r.staff_id=%s");                        params.append(int(staff_id))
    if status:   conds.append("r.status=%s");                          params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT r.*, ps.name as staff_name, ps.employee_code, ps.department, ps.role
            FROM overtime_requests r JOIN punch_staff ps ON ps.id = r.staff_id
            WHERE {' AND '.join(conds)} ORDER BY r.request_date DESC, ps.name
        """, params).fetchall()
    STATUS_LABEL = {'approved': '已核准', 'rejected': '已退回', 'pending': '待審核'}
    wb, ws = _xl_workbook('加班記錄')
    headers = ['員工代碼', '姓名', '部門', '職稱', '加班日期', '開始時間', '結束時間', '時數', '原因', '加班費', '狀態']
    widths  = [10, 10, 12, 12, 12, 8, 8, 7, 24, 10, 8]
    _xl_write_header(ws, headers, widths)
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['staff_name'], r['department'] or '', r['role'] or '',
         str(r['request_date']), str(r['start_time'] or ''), str(r['end_time'] or ''),
         float(r['ot_hours'] or 0), r['reason'] or '',
         float(r['ot_pay'] or 0) if r.get('ot_pay') else '',
         STATUS_LABEL.get(r['status'], r['status'])]
        for r in rows
    ], len(headers), number_cols={8, 10})
    return _xl_response(wb, f'overtime_{month or "all"}.xlsx')


# ── Staff Export ───────────────────────────────────────────────────

@bp.route('/api/export/staff', methods=['GET'])
@login_required
def api_export_staff():
    dept = request.args.get('department', ''); active = request.args.get('active', '1')
    conds, params = ['TRUE'], []
    if active == '1': conds.append("active=TRUE")
    elif active == '0': conds.append("active=FALSE")
    if dept: conds.append("department=%s"); params.append(dept)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT id, employee_code, name, department, role, position_title,
                   salary_type, base_salary, insured_salary,
                   daily_hours, hire_date, birth_date, active, username, line_user_id
            FROM punch_staff WHERE {' AND '.join(conds)} ORDER BY department, name
        """, params).fetchall()
    wb, ws = _xl_workbook('員工資料')
    headers = ['員工代碼', '姓名', '部門', '職稱', '職務', '薪資制度', '底薪', '投保薪資', '每日工時', '到職日', '生日', '狀態', '帳號', 'LINE綁定']
    widths  = [10, 10, 12, 12, 12, 8, 11, 11, 8, 12, 12, 6, 12, 8]
    _xl_write_header(ws, headers, widths)
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['name'], r['department'] or '', r['role'] or '',
         r['position_title'] or '', '時薪制' if r['salary_type'] == 'hourly' else '月薪制',
         float(r['base_salary'] or 0), float(r['insured_salary'] or 0),
         float(r['daily_hours'] or 8),
         str(r['hire_date']) if r['hire_date'] else '',
         str(r['birth_date']) if r['birth_date'] else '',
         '在職' if r['active'] else '離職', r['username'] or '',
         '已綁定' if r['line_user_id'] else '']
        for r in rows
    ], len(headers), number_cols={7, 8, 9})
    return _xl_response(wb, 'staff_list.xlsx')


# ── Expense Export ─────────────────────────────────────────────────

@bp.route('/api/export/expense', methods=['GET'])
@login_required
def api_export_expense():
    month = request.args.get('month', ''); staff_id = request.args.get('staff_id', '')
    status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if month:    conds.append("to_char(ec.expense_date,'YYYY-MM')=%s"); params.append(month)
    if staff_id: conds.append("ec.staff_id=%s");                        params.append(int(staff_id))
    if status:   conds.append("ec.status=%s");                          params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT ec.*, ps.name as staff_name, ps.employee_code, ps.department
            FROM expense_claims ec JOIN punch_staff ps ON ps.id = ec.staff_id
            WHERE {' AND '.join(conds)} ORDER BY ec.expense_date DESC
        """, params).fetchall()
    STATUS_LABEL = {'approved': '已核准', 'rejected': '已退回', 'pending': '待審核'}
    wb, ws = _xl_workbook('費用報帳')
    headers = ['員工代碼', '姓名', '部門', '費用日期', '標題', '金額', '說明', '狀態', '審核人', '申請時間']
    widths  = [10, 10, 12, 12, 24, 11, 30, 8, 10, 16]
    _xl_write_header(ws, headers, widths)
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['staff_name'], r['department'] or '',
         str(r['expense_date']) if r.get('expense_date') else '',
         r['title'] or '', float(r['amount'] or 0), r['note'] or '',
         STATUS_LABEL.get(r['status'], r['status']), r.get('reviewed_by') or '',
         str(r['created_at'])[:16] if r.get('created_at') else '']
        for r in rows
    ], len(headers), number_cols={6})
    return _xl_response(wb, f'expense_{month or "all"}.xlsx')


# ── Leave Balance Export ───────────────────────────────────────────

@bp.route('/api/export/leave-balance', methods=['GET'])
@login_required
def api_export_leave_balance():
    year = request.args.get('year', '') or str(_dt.now(TW_TZ).year)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT lb.*, ps.name as staff_name, ps.employee_code, ps.department,
                   lt.name as leave_type_name, lt.code as leave_code, lt.max_days
            FROM leave_balances lb JOIN punch_staff ps ON ps.id = lb.staff_id
            JOIN leave_types lt ON lt.id = lb.leave_type_id
            WHERE lb.year = %s ORDER BY ps.department, ps.name, lt.sort_order
        """, (int(year),)).fetchall()
    wb, ws = _xl_workbook(f'{year} 請假餘額')
    headers = ['員工代碼', '姓名', '部門', '假別', '假別代碼', '年度上限(天)', '已核准(天)', '剩餘(天)']
    widths  = [10, 10, 12, 12, 8, 11, 11, 10]
    _xl_write_header(ws, headers, widths)
    _xl_write_rows(ws, [
        [r['employee_code'] or '', r['staff_name'], r['department'] or '',
         r['leave_type_name'], r['leave_code'] or '',
         float(r['max_days']) if r['max_days'] is not None else '無限制',
         float(r['used_days'] or 0),
         round(float(r['max_days'] or 0) - float(r['used_days'] or 0), 2)
         if r['max_days'] is not None else '']
        for r in rows
    ], len(headers), number_cols={6, 7, 8})
    return _xl_response(wb, f'leave_balance_{year}.xlsx')


# ── Withholding Tax Export ─────────────────────────────────────────

@bp.route('/api/export/withholding', methods=['GET'])
@require_module('salary')
def api_export_withholding():
    from blueprints.finance import _get_finance_settings
    year = request.args.get('year', str(_date.today().year))
    fmt  = request.args.get('format', 'html')
    fs   = _get_finance_settings()
    company_name    = fs.get('company_name', '')
    company_tax_id  = fs.get('company_tax_id', '')
    company_address = fs.get('company_address', '')

    with get_db() as conn:
        rows = conn.execute("""
            SELECT ps.id, ps.name, ps.national_id, ps.address,
                   COALESCE(SUM(sr.allowance_total), 0) AS gross_salary,
                   COALESCE(SUM(sr.income_tax_withheld), 0) AS tax_withheld,
                   COALESCE(AVG(sr.insured_salary), 0) AS avg_insured
            FROM salary_records sr JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.month LIKE %s AND sr.status='confirmed'
            GROUP BY ps.id, ps.name, ps.national_id, ps.address ORDER BY ps.name
        """, (f'{year}-%',)).fetchall()

    def supp_nhi(gross, insured):
        base = float(gross) - float(insured) * 12
        return max(0, round(base * 0.0211, 0)) if base > 0 else 0

    data = [{'no': i, 'name': r['name'], 'national_id': r['national_id'] or '—',
              'address': r['address'] or '—', 'gross': float(r['gross_salary']),
              'supp_nhi': supp_nhi(r['gross_salary'], r['avg_insured']),
              'tax': float(r['tax_withheld'])}
             for i, r in enumerate(rows, 1)]

    if fmt == 'xlsx':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = f'{year}年扣繳憑單'
        hfill = PatternFill('solid', fgColor='0F1C3A')
        thin  = Border(*[Side(style='thin', color='DDDDDD')]*4)
        hdrs  = ['序號', '姓名', '身分證字號', '地址', '年度薪資合計', '二代健保補充費', '扣繳稅額']
        ws.append(hdrs)
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(1, ci); c.font = Font(bold=True, color='FFFFFF', size=10); c.fill = hfill
            c.alignment = Alignment(horizontal='center', vertical='center'); c.border = thin
        ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 14; ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 16; ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 12
        for d in data:
            ws.append([d['no'], d['name'], d['national_id'], d['address'],
                       d['gross'], d['supp_nhi'], d['tax']])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=withholding_{year}.xlsx'})

    rows_html = ''.join(f"""<tr>
        <td style="text-align:center">{d['no']}</td><td>{d['name']}</td>
        <td style="font-family:monospace">{d['national_id']}</td>
        <td style="font-size:11px">{d['address']}</td>
        <td style="text-align:right;font-family:monospace">{d['gross']:,.0f}</td>
        <td style="text-align:right;font-family:monospace">{d['supp_nhi']:,.0f}</td>
        <td style="text-align:right;font-family:monospace">{d['tax']:,.0f}</td></tr>""" for d in data)
    html = f"""<!DOCTYPE html><html lang="zh-TW"><head><meta charset="UTF-8">
<title>{year}年度薪資扣繳憑單</title>
<style>body{{font-family:'Noto Sans TC',sans-serif;font-size:12px;padding:20px;color:#1e2a45}}
h2{{font-size:16px;font-weight:700;margin-bottom:4px}}
.meta{{font-size:11px;color:#666;margin-bottom:16px}}
table{{width:100%;border-collapse:collapse}}
th{{background:#0f1c3a;color:#fff;padding:7px 10px;font-size:11px;text-align:left}}
td{{padding:6px 10px;border-bottom:1px solid #eee;font-size:12px}}
tr:nth-child(even){{background:#f8f9fb}}
@media print{{button{{display:none}}}}</style></head><body>
<button onclick="window.print()" style="margin-bottom:16px;padding:6px 16px;background:#0f1c3a;color:#fff;border:none;border-radius:4px;cursor:pointer;font-size:12px">列印</button>
<h2>{year} 年度薪資所得扣繳憑單（所得類別 50）</h2>
<div class="meta">扣繳義務人：{company_name}　統一編號：{company_tax_id}　地址：{company_address}　製表日期：{_date.today().isoformat()}</div>
<table><thead><tr><th>#</th><th>員工姓名</th><th>身分證字號</th><th>地址</th>
<th>年度薪資合計(元)</th><th>二代健保補充費(元)</th><th>扣繳稅額(元)</th></tr></thead>
<tbody>{rows_html}</tbody></table></body></html>"""
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


# ── PDF Exports ────────────────────────────────────────────────────

@bp.route('/api/export/attendance/pdf', methods=['GET'])
@login_required
def api_export_attendance_pdf():
    month    = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    staff_id = request.args.get('staff_id', '')
    conds, params = ["TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s"], [month]
    if staff_id:
        conds.append("pr.staff_id=%s"); params.append(int(staff_id))
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT ps.employee_code, ps.name as staff_name, ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   pr.punch_type,
                   to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei', 'HH24:MI') as punch_time,
                   pr.is_manual, pr.location_name
            FROM punch_records pr JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE {' AND '.join(conds)} ORDER BY ps.name, pr.punched_at
        """, params).fetchall()
    PUNCH_LABEL = {'in': '上班', 'out': '下班', 'break_out': '休息開始', 'break_in': '休息結束'}
    headers   = ['員工代碼', '姓名', '部門', '日期', '打卡類型', '時間', '補打', '地點']
    col_widths = [55, 55, 60, 65, 60, 45, 30, 70]
    data = [[r['employee_code'] or '', r['staff_name'], r['department'] or '',
             str(r['work_date']), PUNCH_LABEL.get(r['punch_type'], r['punch_type']),
             r['punch_time'], '是' if r['is_manual'] else '', r['location_name'] or '']
            for r in rows]
    buf = _build_pdf(f'{month} 出勤打卡明細', f'製表：{_date.today().isoformat()}  共 {len(data)} 筆',
                     headers, col_widths, data, landscape=True)
    return _pdf_response(buf, f'attendance_{month}.pdf')


@bp.route('/api/export/attendance-summary/pdf', methods=['GET'])
@login_required
def api_export_attendance_summary_pdf():
    month = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT ps.employee_code, ps.name, ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN to_char(pr.punched_at AT TIME ZONE 'Asia/Taipei','HH24:MI') END) as clock_out,
                   MIN(CASE WHEN pr.punch_type='in'  THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as ci_ts,
                   MAX(CASE WHEN pr.punch_type='out' THEN pr.punched_at AT TIME ZONE 'Asia/Taipei' END) as co_ts,
                   BOOL_OR(pr.is_manual) as has_manual, COUNT(*) as punch_count
            FROM punch_records pr JOIN punch_staff ps ON ps.id = pr.staff_id
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY ps.employee_code, ps.name, ps.department,
                     (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY ps.name, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
        """, (month,)).fetchall()
    headers    = ['員工代碼', '姓名', '部門', '日期', '上班', '下班', '工時(h)', '打卡次數', '含補打']
    col_widths = [55, 55, 60, 65, 45, 45, 50, 55, 40]
    data = []
    for r in rows:
        dur_h = ''
        if r['ci_ts'] and r['co_ts']:
            try:
                ci = r['ci_ts'] if hasattr(r['ci_ts'], 'timestamp') else _dt.fromisoformat(str(r['ci_ts']))
                co = r['co_ts'] if hasattr(r['co_ts'], 'timestamp') else _dt.fromisoformat(str(r['co_ts']))
                dur_h = str(round((co - ci).total_seconds() / 3600, 2))
            except Exception:
                pass
        data.append([r['employee_code'] or '', r['name'], r['department'] or '',
                     str(r['work_date']), r['clock_in'] or '', r['clock_out'] or '',
                     dur_h, str(r['punch_count']), '是' if r['has_manual'] else ''])
    buf = _build_pdf(f'{month} 出勤摘要', f'製表：{_date.today().isoformat()}  共 {len(data)} 筆',
                     headers, col_widths, data, landscape=True)
    return _pdf_response(buf, f'attendance_summary_{month}.pdf')


@bp.route('/api/attendance/anomaly-report/pdf', methods=['GET'])
@login_required
def api_anomaly_report_pdf():
    import calendar as _cal
    month = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    try:
        y, mo = int(month[:4]), int(month[5:7])
    except Exception:
        return jsonify({'error': '月份格式錯誤'}), 400
    with get_db() as conn:
        punch_rows = conn.execute("""
            SELECT ps.id as staff_id, ps.name as staff_name, ps.department,
                   (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date as work_date,
                   MIN(CASE WHEN pr.punch_type='in'  THEN (pr.punched_at AT TIME ZONE 'Asia/Taipei') END) as clock_in,
                   MAX(CASE WHEN pr.punch_type='out' THEN (pr.punched_at AT TIME ZONE 'Asia/Taipei') END) as clock_out,
                   BOOL_OR(pr.punch_type='in')  as has_in,
                   BOOL_OR(pr.punch_type='out') as has_out
            FROM punch_records pr JOIN punch_staff ps ON ps.id=pr.staff_id AND ps.active=TRUE
            WHERE TO_CHAR(pr.punched_at AT TIME ZONE 'Asia/Taipei','YYYY-MM')=%s
            GROUP BY ps.id, ps.name, ps.department, (pr.punched_at AT TIME ZONE 'Asia/Taipei')::date
            ORDER BY work_date, ps.name
        """, (month,)).fetchall()
        shift_rows = conn.execute("""
            SELECT sa.staff_id, sa.shift_date, st.start_time::text, st.end_time::text,
                   ps.name as staff_name, ps.department
            FROM shift_assignments sa JOIN shift_types st ON st.id=sa.shift_type_id
            JOIN punch_staff ps ON ps.id=sa.staff_id AND ps.active=TRUE
            WHERE TO_CHAR(sa.shift_date,'YYYY-MM')=%s
        """, (month,)).fetchall()
        first_day = f"{y}-{mo:02d}-01"
        last_day  = f"{y}-{mo:02d}-{_cal.monthrange(y, mo)[1]:02d}"
        leave_rows = conn.execute("""
            SELECT staff_id, start_date, end_date FROM leave_requests
            WHERE status='approved' AND start_date <= %s AND end_date >= %s
        """, (last_day, first_day)).fetchall()
    from datetime import timedelta as _td
    shift_map = {(r['staff_id'], str(r['shift_date'])): r for r in shift_rows}
    leave_set = set()
    for lr in leave_rows:
        cur = _date.fromisoformat(str(lr['start_date']))
        e   = _date.fromisoformat(str(lr['end_date']))
        while cur <= e:
            leave_set.add((lr['staff_id'], str(cur))); cur += _td(days=1)
    today = _date.today()
    anomalies = []
    for r in punch_rows:
        ds = str(r['work_date']); sid = r['staff_id']
        shift = shift_map.get((sid, ds))
        anomaly_type = ''; detail = ''
        if not r['has_in'] and r['has_out']:
            anomaly_type = '缺上班打卡'; detail = f"僅有下班 {str(r['clock_out'])[11:16]}"
        elif r['has_in'] and not r['has_out']:
            if _date.fromisoformat(ds) < today:
                anomaly_type = '缺下班打卡'; detail = f"上班 {str(r['clock_in'])[11:16]} 無下班"
        elif r['has_in'] and r['has_out'] and shift:
            ci_t = str(r['clock_in'])[11:16]; sh_s = str(shift['start_time'])[:5]
            try:
                late_mins = (int(ci_t[:2])*60+int(ci_t[3:5])) - (int(sh_s[:2])*60+int(sh_s[3:5]))
                if late_mins > 10:
                    anomaly_type = '遲到'; detail = f"應{sh_s}，實{ci_t}(+{late_mins}分)"
            except Exception: pass
            if not anomaly_type:
                co_t = str(r['clock_out'])[11:16]; sh_e = str(shift['end_time'])[:5]
                try:
                    early_mins = (int(sh_e[:2])*60+int(sh_e[3:5])) - (int(co_t[:2])*60+int(co_t[3:5]))
                    if early_mins > 15:
                        anomaly_type = '早退'; detail = f"應{sh_e}，實{co_t}(-{early_mins}分)"
                except Exception: pass
        if anomaly_type:
            anomalies.append([r['staff_name'], r['department'] or '', ds,
                               str(shift['start_time'])[:5] if shift else '—',
                               str(shift['end_time'])[:5]   if shift else '—',
                               str(r['clock_in'])[11:16]  if r['clock_in']  else '—',
                               str(r['clock_out'])[11:16] if r['clock_out'] else '—',
                               anomaly_type, detail])
    punched_set = {(r['staff_id'], str(r['work_date'])) for r in punch_rows}
    for sr in shift_rows:
        ds = str(sr['shift_date']); sid = sr['staff_id']
        if _date.fromisoformat(ds) >= today: continue
        if (sid, ds) in punched_set or (sid, ds) in leave_set: continue
        anomalies.append([sr['staff_name'], sr['department'] or '', ds,
                          str(sr['start_time'])[:5], str(sr['end_time'])[:5],
                          '—', '—', '未打卡', f"排班{str(sr['start_time'])[:5]}~{str(sr['end_time'])[:5]}"])
    anomalies.sort(key=lambda x: (x[2], x[0]))
    headers    = ['姓名', '部門', '日期', '應上班', '應下班', '實際上班', '實際下班', '異常類型', '說明']
    col_widths = [55, 55, 65, 45, 45, 50, 50, 55, 120]
    buf = _build_pdf(f'{month} 出勤異常報告', f'製表：{_date.today().isoformat()}  共 {len(anomalies)} 筆',
                     headers, col_widths, anomalies, landscape=True)
    return _pdf_response(buf, f'anomaly_{month}.pdf')


@bp.route('/api/export/salary/pdf', methods=['GET'])
@login_required
def api_export_salary_pdf():
    month = request.args.get('month', '') or _dt.now(TW_TZ).strftime('%Y-%m')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT sr.*, ps.name as staff_name, ps.employee_code, ps.department, ps.role, ps.salary_type
            FROM salary_records sr JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.month = %s ORDER BY ps.name
        """, (month,)).fetchall()
    headers    = ['代碼', '姓名', '部門', '薪資制', '工作日', '出勤', '請假', '無薪假', '津貼', '扣除', '加班費', '實領', '狀態']
    col_widths = [45, 55, 60, 45, 40, 40, 40, 45, 55, 55, 55, 65, 40]
    data = [[r['employee_code'] or '', r['staff_name'], r['department'] or '',
             '時薪' if r['salary_type'] == 'hourly' else '月薪',
             str(r['work_days'] or 0), str(r['actual_days'] or 0),
             str(r['leave_days'] or 0), str(r['unpaid_days'] or 0),
             f"{float(r['allowance_total'] or 0):,.0f}",
             f"{float(r['deduction_total'] or 0):,.0f}",
             f"{float(r['ot_pay'] or 0):,.0f}",
             f"{float(r['net_pay'] or 0):,.0f}",
             '已確認' if r['status'] == 'confirmed' else '草稿']
            for r in rows]
    buf = _build_pdf(f'{month} 薪資明細', f'製表：{_date.today().isoformat()}  共 {len(data)} 人',
                     headers, col_widths, data, landscape=True)
    return _pdf_response(buf, f'salary_{month}.pdf')


@bp.route('/api/export/leave/pdf', methods=['GET'])
@login_required
def api_export_leave_pdf():
    month = request.args.get('month', ''); year = request.args.get('year', '')
    staff_id = request.args.get('staff_id', ''); status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if status:   conds.append("lr.status=%s");                            params.append(status)
    if month:    conds.append("to_char(lr.start_date,'YYYY-MM')=%s");     params.append(month)
    if year:     conds.append("EXTRACT(YEAR FROM lr.start_date)=%s");      params.append(int(year))
    if staff_id: conds.append("lr.staff_id=%s");                          params.append(int(staff_id))
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT lr.*, ps.name as staff_name, ps.employee_code, ps.department,
                   lt.name as leave_type_name, lt.pay_rate
            FROM leave_requests lr JOIN punch_staff ps ON ps.id = lr.staff_id
            JOIN leave_types lt ON lt.id = lr.leave_type_id
            WHERE {' AND '.join(conds)} ORDER BY lr.start_date, ps.name
        """, params).fetchall()
    PAY_LABEL    = {1.0: '全薪', 0.5: '半薪', 0.0: '無薪'}
    STATUS_LABEL = {'approved': '已核准', 'rejected': '已退回', 'pending': '待審核'}
    headers    = ['員工代碼', '姓名', '部門', '假別', '薪資倍率', '開始日期', '結束日期', '天數', '原因', '代理人', '狀態']
    col_widths = [50, 55, 60, 55, 50, 65, 65, 35, 100, 55, 45]
    data = [[r['employee_code'] or '', r['staff_name'], r['department'] or '',
             r['leave_type_name'], PAY_LABEL.get(float(r['pay_rate']), f"{r['pay_rate']}倍"),
             str(r['start_date']), str(r['end_date']), str(float(r['total_days'])),
             r['reason'] or '', r['substitute_name'] or '',
             STATUS_LABEL.get(r['status'], r['status'])]
            for r in rows]
    label = month or year or 'all'
    buf = _build_pdf(f'{label} 請假記錄', f'製表：{_date.today().isoformat()}  共 {len(data)} 筆',
                     headers, col_widths, data, landscape=True)
    return _pdf_response(buf, f'leave_{label}.pdf')


@bp.route('/api/export/leave-balance/pdf', methods=['GET'])
@login_required
def api_export_leave_balance_pdf():
    year = request.args.get('year', '') or str(_dt.now(TW_TZ).year)
    with get_db() as conn:
        rows = conn.execute("""
            SELECT lb.*, ps.name as staff_name, ps.employee_code, ps.department,
                   lt.name as leave_type_name, lt.code as leave_code, lt.max_days
            FROM leave_balances lb JOIN punch_staff ps ON ps.id = lb.staff_id
            JOIN leave_types lt ON lt.id = lb.leave_type_id
            WHERE lb.year = %s ORDER BY ps.department, ps.name, lt.sort_order
        """, (int(year),)).fetchall()
    headers    = ['員工代碼', '姓名', '部門', '假別', '假別代碼', '年度上限(天)', '已核准(天)', '剩餘(天)']
    col_widths = [55, 55, 65, 65, 50, 65, 65, 55]
    data = [[r['employee_code'] or '', r['staff_name'], r['department'] or '',
             r['leave_type_name'], r['leave_code'] or '',
             str(r['max_days']) if r['max_days'] is not None else '無限制',
             str(float(r['used_days'] or 0)),
             str(round(float(r['max_days'] or 0) - float(r['used_days'] or 0), 2))
             if r['max_days'] is not None else '']
            for r in rows]
    buf = _build_pdf(f'{year} 年假別餘額', f'製表：{_date.today().isoformat()}  共 {len(data)} 筆',
                     headers, col_widths, data)
    return _pdf_response(buf, f'leave_balance_{year}.pdf')


@bp.route('/api/export/overtime/pdf', methods=['GET'])
@login_required
def api_export_overtime_pdf():
    month = request.args.get('month', ''); staff_id = request.args.get('staff_id', '')
    status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if month:    conds.append("to_char(r.request_date,'YYYY-MM')=%s"); params.append(month)
    if staff_id: conds.append("r.staff_id=%s");                        params.append(int(staff_id))
    if status:   conds.append("r.status=%s");                          params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT r.*, ps.name as staff_name, ps.employee_code, ps.department, ps.role
            FROM overtime_requests r JOIN punch_staff ps ON ps.id = r.staff_id
            WHERE {' AND '.join(conds)} ORDER BY r.request_date DESC, ps.name
        """, params).fetchall()
    STATUS_LABEL = {'approved': '已核准', 'rejected': '已退回', 'pending': '待審核'}
    headers    = ['員工代碼', '姓名', '部門', '加班日期', '開始', '結束', '時數', '原因', '加班費', '狀態']
    col_widths = [50, 55, 60, 65, 42, 42, 38, 90, 55, 45]
    data = [[r['employee_code'] or '', r['staff_name'], r['department'] or '',
             str(r['request_date']), str(r['start_time'] or '')[:5], str(r['end_time'] or '')[:5],
             str(float(r['ot_hours'] or 0)), r['reason'] or '',
             f"{float(r['ot_pay'] or 0):,.0f}" if r.get('ot_pay') else '',
             STATUS_LABEL.get(r['status'], r['status'])]
            for r in rows]
    label = month or 'all'
    buf = _build_pdf(f'{label} 加班記錄', f'製表：{_date.today().isoformat()}  共 {len(data)} 筆',
                     headers, col_widths, data, landscape=True)
    return _pdf_response(buf, f'overtime_{label}.pdf')


@bp.route('/api/export/staff/pdf', methods=['GET'])
@login_required
def api_export_staff_pdf():
    dept = request.args.get('department', ''); active = request.args.get('active', '1')
    conds, params = ['TRUE'], []
    if active == '1': conds.append("active=TRUE")
    elif active == '0': conds.append("active=FALSE")
    if dept: conds.append("department=%s"); params.append(dept)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT employee_code, name, department, role, position_title,
                   salary_type, base_salary, insured_salary, hire_date, active
            FROM punch_staff WHERE {' AND '.join(conds)} ORDER BY department, name
        """, params).fetchall()
    headers    = ['員工代碼', '姓名', '部門', '職稱', '職務', '薪資制', '底薪', '投保薪資', '到職日', '狀態']
    col_widths = [55, 55, 65, 65, 65, 45, 60, 65, 65, 40]
    data = [[r['employee_code'] or '', r['name'], r['department'] or '', r['role'] or '',
             r['position_title'] or '', '時薪' if r['salary_type'] == 'hourly' else '月薪',
             f"{float(r['base_salary'] or 0):,.0f}", f"{float(r['insured_salary'] or 0):,.0f}",
             str(r['hire_date']) if r['hire_date'] else '',
             '在職' if r['active'] else '離職']
            for r in rows]
    buf = _build_pdf('員工資料表', f'製表：{_date.today().isoformat()}  共 {len(data)} 人',
                     headers, col_widths, data, landscape=True)
    return _pdf_response(buf, 'staff_list.pdf')


@bp.route('/api/export/training/pdf', methods=['GET'])
@login_required
def api_export_training_pdf():
    staff_id = request.args.get('staff_id', ''); category = request.args.get('category', '')
    conds, params = ['TRUE'], []
    if staff_id: conds.append("tr.staff_id=%s"); params.append(int(staff_id))
    if category: conds.append("tr.category=%s"); params.append(category)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT tr.*, ps.name AS staff_name, ps.department
            FROM training_records tr JOIN punch_staff ps ON tr.staff_id = ps.id
            WHERE {' AND '.join(conds)} ORDER BY tr.expiry_date ASC NULLS LAST, ps.name
        """, params).fetchall()
    CATEGORY_ZH = {'safety': '安全衛生', 'fire': '消防', 'food': '食品衛生',
                   'professional': '專業技能', 'general': '一般訓練'}
    today = _date.today()
    headers    = ['員工姓名', '部門', '課程名稱', '類別', '完訓日期', '到期日', '剩餘天數', '狀態']
    col_widths = [55, 60, 110, 60, 65, 65, 55, 55]
    data = []
    for r in rows:
        days_left = ''; status = '無到期日'
        if r['expiry_date']:
            ed = r['expiry_date'] if hasattr(r['expiry_date'], 'year') else _date.fromisoformat(str(r['expiry_date']))
            days_left = str((ed - today).days)
            if (ed - today).days < 0: status = '已過期'
            elif (ed - today).days <= 60: status = '即將到期'
            else: status = '有效'
        data.append([r['staff_name'], r['department'] or '', r['course_name'],
                     CATEGORY_ZH.get(r['category'], r['category']),
                     str(r['completed_date']) if r['completed_date'] else '',
                     str(r['expiry_date']) if r['expiry_date'] else '',
                     days_left, status])
    buf = _build_pdf('訓練記錄', f'製表：{_date.today().isoformat()}  共 {len(data)} 筆',
                     headers, col_widths, data, landscape=True)
    return _pdf_response(buf, 'training_records.pdf')


@bp.route('/api/export/expense/pdf', methods=['GET'])
@login_required
def api_export_expense_pdf():
    month = request.args.get('month', ''); staff_id = request.args.get('staff_id', '')
    status = request.args.get('status', '')
    conds, params = ['TRUE'], []
    if month:    conds.append("to_char(ec.expense_date,'YYYY-MM')=%s"); params.append(month)
    if staff_id: conds.append("ec.staff_id=%s");                        params.append(int(staff_id))
    if status:   conds.append("ec.status=%s");                          params.append(status)
    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT ec.*, ps.name as staff_name, ps.employee_code, ps.department
            FROM expense_claims ec JOIN punch_staff ps ON ps.id = ec.staff_id
            WHERE {' AND '.join(conds)} ORDER BY ec.expense_date DESC
        """, params).fetchall()
    STATUS_LABEL = {'approved': '已核准', 'rejected': '已退回', 'pending': '待審核'}
    headers    = ['員工代碼', '姓名', '部門', '費用日期', '標題', '金額', '說明', '狀態']
    col_widths = [50, 55, 60, 65, 100, 60, 110, 45]
    data = [[r['employee_code'] or '', r['staff_name'], r['department'] or '',
             str(r['expense_date']) if r.get('expense_date') else '',
             r['title'] or '', f"{float(r['amount'] or 0):,.0f}",
             r['note'] or '', STATUS_LABEL.get(r['status'], r['status'])]
            for r in rows]
    label = month or 'all'
    buf = _build_pdf(f'{label} 費用報帳', f'製表：{_date.today().isoformat()}  共 {len(data)} 筆',
                     headers, col_widths, data, landscape=True)
    return _pdf_response(buf, f'expense_{label}.pdf')


@bp.route('/api/export/withholding/pdf', methods=['GET'])
@require_module('salary')
def api_export_withholding_pdf():
    from blueprints.finance import _get_finance_settings
    year = request.args.get('year', str(_date.today().year))
    fs   = _get_finance_settings()
    company_name = fs.get('company_name', '')
    with get_db() as conn:
        rows = conn.execute("""
            SELECT ps.name, ps.national_id, ps.address,
                   COALESCE(SUM(sr.allowance_total), 0) AS gross_salary,
                   COALESCE(SUM(sr.income_tax_withheld), 0) AS tax_withheld,
                   COALESCE(AVG(sr.insured_salary), 0) AS avg_insured
            FROM salary_records sr JOIN punch_staff ps ON ps.id = sr.staff_id
            WHERE sr.month LIKE %s AND sr.status='confirmed'
            GROUP BY ps.name, ps.national_id, ps.address ORDER BY ps.name
        """, (f'{year}-%',)).fetchall()
    def supp_nhi(gross, insured):
        base = float(gross) - float(insured) * 12
        return max(0, round(base * 0.0211, 0)) if base > 0 else 0
    headers    = ['#', '姓名', '身分證字號', '年度薪資合計', '補充健保費', '扣繳稅額']
    col_widths = [25, 60, 80, 80, 80, 70]
    data = [[str(i), r['name'], r['national_id'] or '—',
             f"{float(r['gross_salary']):,.0f}",
             f"{supp_nhi(r['gross_salary'], r['avg_insured']):,.0f}",
             f"{float(r['tax_withheld']):,.0f}"]
            for i, r in enumerate(rows, 1)]
    subtitle = f"扣繳義務人：{company_name}　製表：{_date.today().isoformat()}  共 {len(data)} 人"
    buf = _build_pdf(f'{year} 年度薪資所得扣繳憑單（所得類別50）', subtitle,
                     headers, col_widths, data)
    return _pdf_response(buf, f'withholding_{year}.pdf')


# ── EDI Exports ────────────────────────────────────────────────────

@bp.route('/api/export/edi/labor-enroll', methods=['GET'])
@require_module('salary')
def api_edi_labor_enroll():
    event_type = request.args.get('event_type', 'in')
    staff_ids  = request.args.get('staff_ids', '')
    event_date = request.args.get('event_date', '')
    cfg        = _get_insurance_settings()
    labor_no   = cfg.get('labor_insurance_no', '').ljust(8)[:8]
    event_code = b'1' if event_type == 'in' else b'2'
    event_roc  = _roc_date(event_date).encode('ascii')
    lines = []
    for s in _get_edi_staff(staff_ids):
        gender_code = b'1' if (s.get('gender') or '').upper() in ('M', '男') else b'2'
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6, '0').encode('ascii')
        lines.append(_edi_bytes(labor_no, 8) + _edi_bytes(s['name'], 20) +
                     _edi_bytes(s.get('national_id', ''), 10) +
                     _roc_date(s.get('birth_date')).encode('ascii') + event_roc + event_code +
                     insured + gender_code + b'00')
    fname = f'labor_{"enroll" if event_type=="in" else "exit"}_{event_date or "date"}.edi'
    return Response(b'\r\n'.join(lines), mimetype='application/octet-stream',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})


@bp.route('/api/export/edi/labor-salary', methods=['GET'])
@require_module('salary')
def api_edi_labor_salary():
    month     = request.args.get('month', '') or _date.today().strftime('%Y-%m')
    staff_ids = request.args.get('staff_ids', '')
    cfg       = _get_insurance_settings()
    labor_no  = cfg.get('labor_insurance_no', '').ljust(8)[:8]
    month_roc = f"{int(month[:4]) - 1911:03d}{month[5:7]}".encode('ascii')
    lines = []
    for s in _get_edi_staff(staff_ids):
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6, '0').encode('ascii')
        lines.append(_edi_bytes(labor_no, 8) + _edi_bytes(s['name'], 20) +
                     _edi_bytes(s.get('national_id', ''), 10) + insured + month_roc)
    return Response(b'\r\n'.join(lines), mimetype='application/octet-stream',
                    headers={'Content-Disposition': f'attachment; filename=labor_salary_{month}.edi'})


@bp.route('/api/export/edi/health-enroll', methods=['GET'])
@require_module('salary')
def api_edi_health_enroll():
    event_type = request.args.get('event_type', 'in')
    staff_ids  = request.args.get('staff_ids', '')
    event_date = request.args.get('event_date', '')
    cfg        = _get_insurance_settings()
    health_no  = cfg.get('health_insurance_no', '').ljust(10)[:10]
    event_code = b'1' if event_type == 'in' else b'2'
    event_roc  = _roc_date(event_date).encode('ascii')
    lines = []
    for s in _get_edi_staff(staff_ids):
        gender_code = b'1' if (s.get('gender') or '').upper() in ('M', '男') else b'2'
        insured = str(int(float(s.get('insured_salary') or 0))).rjust(6, '0').encode('ascii')
        lines.append(_edi_bytes(health_no, 10) + _edi_bytes(s['name'], 20) +
                     _edi_bytes(s.get('national_id', ''), 10) +
                     _roc_date(s.get('birth_date')).encode('ascii') + event_roc + event_code +
                     insured + gender_code)
    fname = f'health_{"enroll" if event_type=="in" else "exit"}_{event_date or "date"}.edi'
    return Response(b'\r\n'.join(lines), mimetype='application/octet-stream',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})
