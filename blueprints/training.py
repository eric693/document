"""
blueprints/training.py — 教育訓練追蹤模組
"""
from datetime import date, datetime as _dtm

from flask import Blueprint, request, jsonify

from auth import login_required, require_module
from config import TW_TZ
from db import get_db
from blueprints.exports import _xl_workbook, _xl_write_header, _xl_response

bp = Blueprint('training', __name__)


# ─── DB init ─────────────────────────────────────────────────────────────────

def init_training_db():
    try:
        with get_db() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS training_records (
                    id              SERIAL PRIMARY KEY,
                    staff_id        INT REFERENCES punch_staff(id) ON DELETE CASCADE,
                    course_name     TEXT NOT NULL,
                    category        TEXT NOT NULL DEFAULT 'general',
                    completed_date  DATE,
                    expiry_date     DATE,
                    certificate_no  TEXT DEFAULT '',
                    note            TEXT DEFAULT '',
                    created_at      TIMESTAMPTZ DEFAULT NOW(),
                    updated_at      TIMESTAMPTZ DEFAULT NOW()
                )
            """)
    except Exception as e:
        print(f"[training_init] {e}")


TRAINING_CATEGORIES = {
    'food_safety':  '食品安全',
    'fire_safety':  '消防安全',
    'first_aid':    '急救訓練',
    'hygiene':      '衛生管理',
    'service':      '服務禮儀',
    'equipment':    '設備操作',
    'general':      '一般訓練',
    'other':        '其他',
}


# ─── Routes ──────────────────────────────────────────────────────────────────

@bp.route('/api/training/records', methods=['GET'])
@require_module('training')
def api_training_list():
    from datetime import datetime as _dt_tr
    staff_id  = request.args.get('staff_id')
    category  = request.args.get('category', '')
    expiring  = request.args.get('expiring')
    expired   = request.args.get('expired')

    sql = """
        SELECT tr.*, ps.name AS staff_name, ps.department
        FROM training_records tr
        JOIN punch_staff ps ON tr.staff_id = ps.id
        WHERE 1=1
    """
    params = []
    if staff_id:
        sql += " AND tr.staff_id = %s"; params.append(int(staff_id))
    if category:
        sql += " AND tr.category = %s"; params.append(category)
    if expiring:
        days = int(expiring)
        sql += " AND tr.expiry_date IS NOT NULL AND tr.expiry_date <= CURRENT_DATE + INTERVAL '%s days' AND tr.expiry_date >= CURRENT_DATE"
        params.append(days)
    if expired == '1':
        sql += " AND tr.expiry_date IS NOT NULL AND tr.expiry_date < CURRENT_DATE"
    sql += " ORDER BY tr.expiry_date ASC NULLS LAST, tr.completed_date DESC"

    with get_db() as conn:
        rows = conn.execute(sql, params).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        for k in ('completed_date', 'expiry_date', 'created_at', 'updated_at'):
            if d.get(k): d[k] = str(d[k])
        today = _dtm.now(TW_TZ).date()
        if d.get('expiry_date'):
            ed = _dt_tr.strptime(d['expiry_date'], '%Y-%m-%d').date()
            days_left = (ed - today).days
            d['days_left'] = days_left
            d['status'] = 'expired' if days_left < 0 else 'expiring_soon' if days_left <= 60 else 'valid'
        else:
            d['days_left'] = None
            d['status'] = 'no_expiry'
        result.append(d)
    return jsonify(result)


@bp.route('/api/training/records', methods=['POST'])
@require_module('training')
def api_training_create():
    b = request.get_json(force=True) or {}
    staff_id       = b.get('staff_id')
    course_name    = (b.get('course_name') or '').strip()
    category       = b.get('category', 'general')
    completed_date = b.get('completed_date') or None
    expiry_date    = b.get('expiry_date') or None
    certificate_no = (b.get('certificate_no') or '').strip()
    note           = (b.get('note') or '').strip()
    if not staff_id or not course_name:
        return jsonify({'error': '缺少必填欄位'}), 400
    with get_db() as conn:
        row = conn.execute("""
            INSERT INTO training_records
              (staff_id, course_name, category, completed_date, expiry_date, certificate_no, note)
            VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id
        """, (staff_id, course_name, category, completed_date, expiry_date, certificate_no, note)).fetchone()
    return jsonify({'ok': True, 'id': row['id']})


@bp.route('/api/training/records/<int:rid>', methods=['PUT'])
@require_module('training')
def api_training_update(rid):
    b = request.get_json(force=True) or {}
    with get_db() as conn:
        conn.execute("""
            UPDATE training_records SET
              course_name=%s, category=%s, completed_date=%s, expiry_date=%s,
              certificate_no=%s, note=%s, updated_at=NOW()
            WHERE id=%s
        """, (
            b.get('course_name'), b.get('category', 'general'),
            b.get('completed_date') or None, b.get('expiry_date') or None,
            b.get('certificate_no', ''), b.get('note', ''), rid
        ))
    return jsonify({'ok': True})


@bp.route('/api/training/records/<int:rid>', methods=['DELETE'])
@require_module('training')
def api_training_delete(rid):
    with get_db() as conn:
        conn.execute("DELETE FROM training_records WHERE id=%s", (rid,))
    return jsonify({'ok': True})


@bp.route('/api/training/summary', methods=['GET'])
@require_module('training')
def api_training_summary():
    """每位員工的訓練狀況摘要"""
    with get_db() as conn:
        staff_all = conn.execute(
            "SELECT id, name, department FROM punch_staff WHERE active=TRUE ORDER BY name"
        ).fetchall()
        records = conn.execute("""
            SELECT staff_id, category, expiry_date,
                   CASE
                     WHEN expiry_date IS NULL THEN 'no_expiry'
                     WHEN expiry_date < CURRENT_DATE THEN 'expired'
                     WHEN expiry_date <= CURRENT_DATE + INTERVAL '60 days' THEN 'expiring_soon'
                     ELSE 'valid'
                   END AS status
            FROM training_records
        """).fetchall()
    from collections import defaultdict
    by_staff = defaultdict(list)
    for r in records:
        by_staff[r['staff_id']].append(dict(r))

    result = []
    for s in staff_all:
        recs = by_staff[s['id']]
        result.append({
            'id': s['id'], 'name': s['name'], 'department': s['department'],
            'total': len(recs),
            'valid': sum(1 for r in recs if r['status'] in ('valid', 'no_expiry')),
            'expiring_soon': sum(1 for r in recs if r['status'] == 'expiring_soon'),
            'expired': sum(1 for r in recs if r['status'] == 'expired'),
        })
    return jsonify(result)


@bp.route('/api/export/training', methods=['GET'])
@require_module('training')
def api_export_training():
    """匯出訓練記錄 Excel"""
    staff_id = request.args.get('staff_id', '')
    category = request.args.get('category', '')

    conds, params = ['TRUE'], []
    if staff_id: conds.append("tr.staff_id=%s"); params.append(int(staff_id))
    if category: conds.append("tr.category=%s"); params.append(category)

    with get_db() as conn:
        rows = conn.execute(f"""
            SELECT tr.*, ps.name AS staff_name, ps.department
            FROM training_records tr
            JOIN punch_staff ps ON tr.staff_id = ps.id
            WHERE {' AND '.join(conds)}
            ORDER BY tr.expiry_date ASC NULLS LAST, ps.name
        """, params).fetchall()

    today = _dtm.now(TW_TZ).date()
    CATEGORY_ZH = {'safety':'安全衛生','fire':'消防','food':'食品衛生',
                   'professional':'專業技能','general':'一般訓練'}
    for k, v in TRAINING_CATEGORIES.items():
        CATEGORY_ZH[k] = v

    from openpyxl.styles import PatternFill, Alignment, Border, Side
    wb, ws = _xl_workbook('訓練記錄')
    headers = ['員工姓名','案場','課程名稱','類別','完訓日期','到期日','證書號碼','剩餘天數','狀態','備註']
    widths  = [10, 12, 24, 10, 12, 12, 16, 9, 10, 20]
    _xl_write_header(ws, headers, widths)

    warn_fill = PatternFill('solid', fgColor='FFF3CD')
    err_fill  = PatternFill('solid', fgColor='FDECEA')
    thin = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin',  color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
    )
    even_fill = PatternFill('solid', fgColor='F4F6FA')
    center    = Alignment(horizontal='center', vertical='center')
    left_al   = Alignment(horizontal='left', vertical='center')

    data = []
    row_colors = []
    for r in rows:
        expiry = str(r['expiry_date']) if r['expiry_date'] else ''
        days_left = ''
        status = '無到期日'
        color = None
        if r['expiry_date']:
            ed = r['expiry_date'] if hasattr(r['expiry_date'], 'year') else date.fromisoformat(str(r['expiry_date']))
            days_left = (ed - today).days
            if days_left < 0:
                status = '已過期'; color = 'err'
            elif days_left <= 60:
                status = '即將到期'; color = 'warn'
            else:
                status = '有效'
        data.append([r['staff_name'], r['department'] or '',
                     r['course_name'], CATEGORY_ZH.get(r['category'], r['category']),
                     str(r['completed_date']) if r['completed_date'] else '',
                     expiry, r['certificate_no'] or '',
                     days_left, status, r['note'] or ''])
        row_colors.append(color)

    for ri, (row_vals, color) in enumerate(zip(data, row_colors), 2):
        fill = err_fill if color == 'err' else warn_fill if color == 'warn' else (even_fill if ri % 2 == 0 else None)
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            if fill: cell.fill = fill
            cell.border = thin
            cell.alignment = center if isinstance(v, (int, float, type(None))) else left_al
    return _xl_response(wb, 'training_records.xlsx')
