"""
blueprints/export_utils.py — Excel 匯出輔助函式（跨模組共用）
"""
import io
from flask import Response


def _xl_workbook():
    """建立帶有預設字型的 openpyxl Workbook"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    return wb


def _xl_write_header(ws, headers, row=1):
    """將 headers 清單寫入指定行，套用粗體與淺灰底色"""
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill('solid', fgColor='D9D9D9')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _xl_write_rows(ws, rows, start_row=2):
    """將二維資料（list of list/tuple）寫入工作表"""
    for r_idx, row_data in enumerate(rows, start_row):
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def _xl_response(wb, filename):
    """將 openpyxl Workbook 包成 Flask Response 回傳下載"""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
