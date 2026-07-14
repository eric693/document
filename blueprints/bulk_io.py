"""
blueprints/bulk_io.py — 整批匯入／匯出（員工名單、文件管理狀態）

新增模組，不動既有程式。提供：
  員工名單  GET  /api/bulk/staff/template     下載匯入範本
            GET  /api/bulk/staff/export       匯出現有員工（格式與範本相同，可回填再匯入）
            POST /api/bulk/staff/import        上傳 Excel 批次新增／更新
  文件管理  GET  /api/bulk/documents/template  下載矩陣範本（員工已帶入、文件欄位留空）
            GET  /api/bulk/documents/export    匯出收件狀態矩陣
            POST /api/bulk/documents/import    上傳矩陣批次登記狀態

匯入採「以員工編號→帳號→姓名」比對；空白儲存格＝不變更（避免覆寫既有資料）。
"""
import io
import re
import secrets
from datetime import datetime

from flask import Blueprint, request, jsonify, session, Response

import psycopg
from psycopg.types.json import Json
from auth import require_module
from db import get_db, hash_password
from blueprints.audit import log_action

bp = Blueprint('bulk_io', __name__)


# ─── 共用工具 ────────────────────────────────────────────────────────────────

def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _style_header(ws, ncols, row=1):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill('solid', fgColor='1F3A5F')
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _cell_str(v):
    """儲存格轉為乾淨字串（去除多餘空白、處理 datetime）"""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if hasattr(v, 'isoformat'):   # date
        return v.isoformat()
    return str(v).strip()


def _parse_date(s):
    """接受 YYYY-MM-DD / YYYY/MM/DD / 民國年，回傳 ISO 字串或 None"""
    s = _cell_str(s)
    if not s:
        return None
    if isinstance(s, str) and 'T' in s:      # datetime isoformat
        s = s.split('T')[0]
    parts = re.split(r'[/\-\.]', s)
    if len(parts) != 3:
        return None
    try:
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
    except ValueError:
        return None
    if y < 1911:          # 民國年
        y += 1911
    if not (1 <= m <= 12 and 1 <= d <= 31):
        return None
    return f'{y:04d}-{m:02d}-{d:02d}'


def _read_upload():
    """讀取上傳的 .xlsx，回傳 (rows, error)。rows 為 list[list]（含表頭）"""
    file = request.files.get('file')
    if not file or not file.filename:
        return None, '請上傳 Excel 檔案（.xlsx）'
    if not file.filename.lower().endswith('.xlsx'):
        return None, '僅支援 .xlsx 格式，請另存新檔後再上傳'
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True, read_only=True)
        ws = wb.active
        rows = [[_cell_str(c) for c in row] for row in ws.iter_rows(values_only=True)]
    except Exception as e:
        return None, f'檔案讀取失敗：{e}'
    # 去掉整列皆空的資料
    rows = [r for r in rows if any(c for c in r)]
    if len(rows) < 2:
        return None, '檔案沒有可匯入的資料列'
    return rows, None


# ═════════════════════════════════════════════════════════════════════════════
#  員工名單
# ═════════════════════════════════════════════════════════════════════════════

# (表頭中文, punch_staff 欄位, 類型)  類型：text / date / active / password
STAFF_COLUMNS = [
    ('員工編號',   'employee_code',  'text'),
    ('姓名',       'name',           'text'),
    ('帳號',       'username',       'text'),
    ('密碼',       'password',       'password'),
    ('案場',       'department',     'text'),
    ('職稱',       'position_title', 'text'),
    ('公司',       'company',        'text'),
    ('身分證字號', 'national_id',    'text'),
    ('電話',       'phone',          'text'),
    ('緊急聯絡人', 'emergency_contact', 'text'),
    ('前科',       'criminal_record', 'text'),
    ('備註',       'staff_note',     'text'),
    ('性別',       'gender',         'text'),
    ('生日',       'birth_date',     'date'),
    ('到職日',     'hire_date',      'date'),
    ('地址',       'address',        'text'),
    ('銀行代碼',   'bank_code',      'text'),
    ('銀行',       'bank_name',      'text'),
    ('分行',       'bank_branch',    'text'),
    ('匯款帳號',   'bank_account',   'text'),
    ('戶名',       'account_holder', 'text'),
    ('狀態',       'active',         'active'),
]
_STAFF_EXPORT_FIELDS = [c[1] for c in STAFF_COLUMNS if c[1] != 'password']


def _active_field_defs(conn):
    """啟用中的自訂欄位定義（匯入匯出動態欄）"""
    try:
        return [dict(r) for r in conn.execute(
            "SELECT name, field_type FROM staff_field_defs WHERE active=TRUE ORDER BY sort_order, id"
        ).fetchall()]
    except Exception:
        return []


def _staff_wb(rows_data, extra_headers=None):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '員工名單'
    headers = [c[0] for c in STAFF_COLUMNS] + list(extra_headers or [])
    ws.append(headers)
    for r in rows_data:
        ws.append(r)
    _style_header(ws, len(headers))
    widths = [12, 12, 14, 12, 12, 12, 12, 16, 13, 12, 6, 20, 6, 12, 12, 24, 10, 12, 12, 16, 12, 8]
    widths += [14] * len(extra_headers or [])
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    # 說明分頁
    note = wb.create_sheet('填寫說明')
    for line in [
        ['整批匯入員工說明'],
        [''],
        ['1. 比對順序：員工編號 → 帳號 → 姓名。找到既有員工則「更新」，否則「新增」。'],
        ['2. 空白儲存格＝不變更該欄位（更新時不會清空既有資料）。'],
        ['3. 新增員工時「姓名」為必填；帳號留空會自動以員工編號或系統代碼產生。'],
        ['4. 新增員工時「密碼」留空會自動產生 8 碼亂數，匯入結果會列出帳號密碼。'],
        ['5. 日期可用 2024-01-05、2024/1/5 或民國年（113/01/05）。'],
        ['6. 狀態欄填「在職」或「離職」，留空視為在職。'],
        ['7. 「密碼」欄只用於匯入，匯出時一律留空。'],
    ]:
        note.append(line)
    note.column_dimensions['A'].width = 80
    return wb


@bp.route('/api/bulk/staff/template', methods=['GET'])
@require_module('punch')
def staff_template():
    with get_db() as conn:
        defs = _active_field_defs(conn)
    example = [
        'A001', '王小明', 'wang.ming', '', '一號案場', '店員', '永興公司',
        'A123456789', '0912345678', '王大明 0987654321', '無', '',
        '男', '1990-05-20', '2024-01-05', '台北市信義區…',
        '822', '中國信託', '信義分行', '1234567890123', '王小明', '在職',
    ] + [''] * len(defs)
    return _xlsx_response(_staff_wb([example], [d['name'] for d in defs]),
                          'staff_import_template.xlsx')


@bp.route('/api/bulk/staff/export', methods=['GET'])
@require_module('punch')
def staff_export():
    active = request.args.get('active', '')
    dept   = request.args.get('department', '')
    conds, params = ['TRUE'], []
    if active == '1':   conds.append('active=TRUE')
    elif active == '0': conds.append('active=FALSE')
    if dept:
        conds.append('department=%s'); params.append(dept)
    with get_db() as conn:
        # 只取匯出需要的欄位（排除照片等大欄位）
        staff = conn.execute(
            "SELECT employee_code, name, username, department, position_title, company, "
            "national_id, phone, emergency_contact, criminal_record, staff_note, gender, birth_date, hire_date, address, "
            "bank_code, bank_name, bank_branch, bank_account, account_holder, active, custom_fields "
            f"FROM punch_staff WHERE {' AND '.join(conds)} "
            "ORDER BY department, sort_order, id", params
        ).fetchall()
        defs = _active_field_defs(conn)
    rows_data = []
    for s in staff:
        row = []
        for zh, field, typ in STAFF_COLUMNS:
            if typ == 'password':
                row.append('')
            elif typ == 'active':
                row.append('在職' if s['active'] else '離職')
            else:
                row.append(_cell_str(s[field]))
        cf = s['custom_fields'] or {}
        row += [_cell_str(cf.get(d['name'])) for d in defs]
        rows_data.append(row)
    return _xlsx_response(_staff_wb(rows_data, [d['name'] for d in defs]), 'staff_list.xlsx')


def _gen_password():
    # 8 碼、避開易混淆字元
    alphabet = '23456789ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnpqrstuvwxyz'
    return ''.join(secrets.choice(alphabet) for _ in range(8))


@bp.route('/api/bulk/staff/import', methods=['POST'])
@require_module('punch')
def staff_import():
    rows, err = _read_upload()
    if err:
        return jsonify({'error': err}), 400

    header = rows[0]
    # 建立「欄位 index」對照（依表頭中文比對，容許順序不同、缺欄位）
    zh_to_field = {zh: (field, typ) for zh, field, typ in STAFF_COLUMNS}
    zh_to_field['部門'] = ('department', 'text')   # 相容舊版表頭
    col_idx = {}   # field -> column index
    col_type = {}
    for i, h in enumerate(header):
        h = h.strip()
        if h in zh_to_field:
            field, typ = zh_to_field[h]
            col_idx[field] = i
            col_type[field] = typ
    if 'name' not in col_idx:
        return jsonify({'error': '找不到「姓名」欄位，請使用下載的範本格式'}), 400

    preview = request.args.get('preview') == '1'
    created = updated = skipped = 0
    errors, credentials = [], []
    plan_creates, plan_updates = [], []

    with get_db() as conn:
        # 自訂欄位：表頭比對啟用中的欄位定義（固定欄位優先）
        def_names = {d['name'] for d in _active_field_defs(conn)}
        cf_idx = {}   # 欄位名 -> column index
        for i, h in enumerate(header):
            h = h.strip()
            if h in def_names and h not in zh_to_field:
                cf_idx[h] = i

        for rownum, r in enumerate(rows[1:], start=2):
            def val(field):
                i = col_idx.get(field)
                return r[i].strip() if (i is not None and i < len(r)) else ''

            cf = {}
            for cfname, ci in cf_idx.items():
                v = r[ci].strip() if ci < len(r) else ''
                if v != '':
                    cf[cfname] = v

            name = val('name')
            emp  = val('employee_code')
            user = val('username')
            if not name and not emp and not user:
                continue   # 空列

            # 比對既有員工
            match = None
            if emp:
                match = conn.execute(
                    'SELECT * FROM punch_staff WHERE employee_code=%s', (emp,)).fetchone()
            if not match and user:
                match = conn.execute(
                    'SELECT * FROM punch_staff WHERE username=%s', (user,)).fetchone()
            if not match and name:
                match = conn.execute(
                    'SELECT * FROM punch_staff WHERE name=%s', (name,)).fetchone()

            # 收集要寫入的欄位（僅非空欄位）
            fields = {}
            for field in col_idx:
                if field in ('password',):
                    continue
                raw = val(field)
                if col_type[field] == 'active':
                    if raw:
                        fields['active'] = raw not in ('離職', '停用', '否', '0', 'N', 'false')
                elif col_type[field] == 'date':
                    d = _parse_date(raw)
                    if d:
                        fields[field] = d
                else:
                    if raw != '':
                        fields[field] = raw

            # 每列包在 savepoint，單列失敗不影響其他列
            try:
                if match:
                    if not fields and not cf:
                        skipped += 1
                        continue
                    if preview:
                        changes = {}
                        for k, v in fields.items():
                            cur = match.get(k)
                            cur_s = ('在職' if cur else '離職') if k == 'active' else _cell_str(cur)
                            new_s = ('在職' if v else '離職') if k == 'active' else str(v)
                            if cur_s != new_s:
                                changes[k] = [cur_s, new_s]
                        mcf = match.get('custom_fields') or {}
                        for k, v in cf.items():
                            if str(mcf.get(k, '')) != v:
                                changes[k] = [str(mcf.get(k, '')), v]
                        if changes:
                            plan_updates.append({'name': match['name'], 'changes': changes})
                            updated += 1
                        else:
                            skipped += 1
                        continue
                    with conn.transaction():
                        if fields:
                            sets = ', '.join(f'{k}=%s' for k in fields)
                            conn.execute(
                                f'UPDATE punch_staff SET {sets} WHERE id=%s',
                                list(fields.values()) + [match['id']])
                        if cf:
                            conn.execute(
                                "UPDATE punch_staff SET custom_fields = "
                                "COALESCE(custom_fields,'{}'::jsonb) || %s::jsonb WHERE id=%s",
                                (Json(cf), match['id']))
                    updated += 1
                else:
                    if not name:
                        errors.append(f'第 {rownum} 列：新增需要「姓名」，已略過')
                        continue
                    username = user or emp or f'u{secrets.token_hex(3)}'
                    if conn.execute('SELECT 1 FROM punch_staff WHERE username=%s',
                                    (username,)).fetchone():
                        errors.append(f'第 {rownum} 列：帳號「{username}」已存在，已略過')
                        continue
                    pw = val('password')
                    gen = False
                    if not pw:
                        pw = _gen_password(); gen = True
                    elif len(pw) < 8:
                        errors.append(f'第 {rownum} 列：密碼少於 8 碼，已略過')
                        continue
                    fields.pop('active', None)   # 新增預設在職
                    if preview:
                        plan_creates.append({'name': name, 'username': username,
                                             'auto_password': gen})
                        created += 1
                        continue
                    cols = ['name', 'username', 'password_hash', 'password_plain', 'custom_fields'] + \
                           [k for k in fields if k != 'name']
                    vals = [name, username, hash_password(pw), pw, Json(cf)] + \
                           [fields[k] for k in fields if k != 'name']
                    placeholders = ','.join(['%s'] * len(cols))
                    with conn.transaction():
                        conn.execute(
                            f'INSERT INTO punch_staff ({",".join(cols)}) VALUES ({placeholders})',
                            vals)
                    created += 1
                    if gen:
                        credentials.append({'name': name, 'username': username, 'password': pw})
            except psycopg.errors.UniqueViolation:
                errors.append(f'第 {rownum} 列：{name} 資料重複（姓名／帳號），已略過')
            except Exception as e:
                errors.append(f'第 {rownum} 列：{name} 匯入失敗（{e}）')

    if preview:
        return jsonify({
            'preview': True, 'created': created, 'updated': updated, 'skipped': skipped,
            'errors': errors, 'creates': plan_creates, 'updates': plan_updates,
        })
    log_action('整批匯入員工', '', f'新增 {created}、更新 {updated}、略過 {skipped}、錯誤 {len(errors)}')
    return jsonify({
        'created': created, 'updated': updated, 'skipped': skipped,
        'errors': errors, 'credentials': credentials,
    })


# ═════════════════════════════════════════════════════════════════════════════
#  文件管理狀態矩陣
# ═════════════════════════════════════════════════════════════════════════════

_DOC_ID_HEADERS = ['員工編號', '姓名', '部門']


def _doc_matrix_wb(include_status):
    """建立矩陣 Workbook。include_status=False 時文件欄留空（範本用）"""
    from blueprints.documents import _build_matrix
    import openpyxl
    with get_db() as conn:
        types, staff_rows = _build_matrix(conn)
        staff_meta = {s['id']: s for s in conn.execute(
            'SELECT id, employee_code FROM punch_staff').fetchall()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '文件收件狀態'
    headers = list(_DOC_ID_HEADERS) + [t['name'] for t in types]
    ws.append(headers)
    for srow in staff_rows:
        emp = _cell_str(staff_meta.get(srow['id'], {}).get('employee_code')) \
            if srow['id'] in staff_meta else ''
        line = [emp, srow['name'], srow['department']]
        for t in types:
            if not include_status:
                line.append('')
                continue
            cell = srow['items'].get(str(t['id']), {})
            st = cell.get('status')
            if st == 'received':
                line.append(cell.get('content') or '✓')
            elif st == 'na':
                line.append('免')
            else:
                line.append('✗')
        ws.append(line)
    _style_header(ws, len(headers))
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    for i in range(len(_DOC_ID_HEADERS) + 1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12

    note = wb.create_sheet('填寫說明')
    for line in [
        ['整批匯入文件收件狀態說明'],
        [''],
        ['1. 前三欄用來比對員工（員工編號優先，其次姓名），請勿更動。'],
        ['2. 文件欄位填法：'],
        ['   ✓ 或 已收 或 o／v／是／有 → 標記為「已收」'],
        ['   ✗ 或 缺 或 x／否 → 標記為「缺件」'],
        ['   免 或 免附 → 標記為「免附」'],
        ['   清 或 clear → 清除手動記錄，回復系統自動判定'],
        ['   直接填文字（例如電話號碼）→ 標記為已收並記錄該內容'],
        ['   留空 → 不變更該格狀態'],
        ['3. 只會處理「手動登記」的文件欄位；綁定員工資料的欄位（姓名、照片、電話等）'],
        ['   由員工資料自動判定，匯入時會忽略。未知欄位也會被忽略。'],
    ]:
        note.append(line)
    note.column_dimensions['A'].width = 70
    return wb


@bp.route('/api/bulk/documents/template', methods=['GET'])
@require_module('docs')
def documents_template():
    return _xlsx_response(_doc_matrix_wb(include_status=False),
                          'documents_matrix_template.xlsx')


@bp.route('/api/bulk/documents/export', methods=['GET'])
@require_module('docs')
def documents_export():
    return _xlsx_response(_doc_matrix_wb(include_status=True),
                          'documents_matrix.xlsx')


def _parse_doc_cell(s):
    """回傳 (status, content) 或 None（不變更）"""
    s = _cell_str(s)
    if s == '':
        return None
    low = s.lower()
    if low in ('✓', 'v', 'o', 'y', 'yes', '是', '有', '已收', 'received', '1', 'true', '✔'):
        return ('received', '')
    if low in ('✗', 'x', '缺', '缺件', 'missing', 'no', '否', '無', '0', 'false', '✘'):
        return ('missing', '')
    if low in ('免', '免附', 'na', 'n/a', '不適用'):
        return ('na', '')
    if low in ('清', 'clear', '回復自動', '自動'):
        return ('clear', '')
    return ('received', s)   # 其他文字視為已收並記錄內容


@bp.route('/api/bulk/documents/import', methods=['POST'])
@require_module('docs')
def documents_import():
    rows, err = _read_upload()
    if err:
        return jsonify({'error': err}), 400

    header = [h.strip() for h in rows[0]]
    with get_db() as conn:
        # 僅接受「手動登記」項目；綁定員工欄位的項目由員工資料自動判定，
        # 匯入時忽略，避免建立手動記錄蓋住自動狀態
        types = conn.execute(
            "SELECT id, name FROM document_types "
            "WHERE active=TRUE AND (staff_field='' OR staff_field IS NULL)").fetchall()
        name_to_type = {t['name']: t['id'] for t in types}

        # 前 len(_DOC_ID_HEADERS) 欄為身分識別欄，其後才是文件欄位。
        # （文件項目中也可能有「姓名」「員工編號」，用位置區隔避免撞名）
        nid = len(_DOC_ID_HEADERS)
        emp_idx = name_idx = None
        for i in range(min(nid, len(header))):
            if header[i] == '員工編號' and emp_idx is None:
                emp_idx = i
            elif header[i] == '姓名' and name_idx is None:
                name_idx = i
        doc_cols = {}   # col_index -> doc_type_id
        for i in range(nid, len(header)):
            if header[i] in name_to_type:
                doc_cols[i] = name_to_type[header[i]]
        if name_idx is None and emp_idx is None:
            return jsonify({'error': '找不到「員工編號」或「姓名」欄位'}), 400
        if not doc_cols:
            return jsonify({'error': '找不到任何對應的文件項目欄位'}), 400

        updated_cells = cleared = 0
        rows_matched = 0
        errors = []
        who = session.get('admin_display_name', '管理員')

        for rownum, r in enumerate(rows[1:], start=2):
            emp  = r[emp_idx].strip()  if emp_idx  is not None and emp_idx  < len(r) else ''
            name = r[name_idx].strip() if name_idx is not None and name_idx < len(r) else ''
            if not emp and not name:
                continue
            staff = None
            if emp:
                staff = conn.execute(
                    'SELECT id FROM punch_staff WHERE employee_code=%s', (emp,)).fetchone()
            if not staff and name:
                staff = conn.execute(
                    'SELECT id FROM punch_staff WHERE name=%s', (name,)).fetchone()
            if not staff:
                errors.append(f'第 {rownum} 列：找不到員工「{name or emp}」，已略過')
                continue
            sid = staff['id']
            row_updates = row_clears = 0
            try:
                with conn.transaction():
                    for ci, tid in doc_cols.items():
                        if ci >= len(r):
                            continue
                        parsed = _parse_doc_cell(r[ci])
                        if parsed is None:
                            continue
                        status, content = parsed
                        if status == 'clear':
                            conn.execute(
                                'DELETE FROM staff_documents '
                                'WHERE staff_id=%s AND doc_type_id=%s', (sid, tid))
                            row_clears += 1
                            continue
                        conn.execute(
                            """INSERT INTO staff_documents
                                 (staff_id, doc_type_id, status, content, updated_by)
                               VALUES (%s,%s,%s,%s,%s)
                               ON CONFLICT (staff_id, doc_type_id) DO UPDATE SET
                                 status=EXCLUDED.status, content=EXCLUDED.content,
                                 updated_by=EXCLUDED.updated_by, updated_at=NOW()""",
                            (sid, tid, status, content, who))
                        row_updates += 1
            except Exception as e:
                errors.append(f'第 {rownum} 列：{name or emp} 匯入失敗（{e}）')
                continue
            rows_matched += 1
            updated_cells += row_updates
            cleared += row_clears

    log_action('整批匯入文件狀態', '', f'{rows_matched} 位員工、更新 {updated_cells} 格')
    return jsonify({
        'rows_matched': rows_matched, 'updated_cells': updated_cells,
        'cleared': cleared, 'errors': errors,
    })
